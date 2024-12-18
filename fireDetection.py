import threading
import cv2
import pygame
import smtplib
from email.message import EmailMessage
import ssl
import openpyxl
from openpyxl import Workbook
from datetime import datetime
from twilio.rest import Client


account_sid = 'your sid'
auth_token = 'your token'


client = Client(account_sid, auth_token)
def call():
    call = client.calls.create(
                            url="url for twillio account",
                            to='number of the user',
                            from_='number by twillio'
                        )

    print(call.sid)
client = Client(account_sid, auth_token)


def sendsms():
    message = client.messages.create(
        body='Fire has initiated in the building. Please exit the building from the nearest exit.',
        from_='number by twilio with country code',
        to='number of the user with country code',
    )
    print("SMS message sent successfully. SID:", message.sid)

    

fire_cascade = cv2.CascadeClassifier('fire_detection_cascade_model.xml') 
vid = cv2.VideoCapture(0)
runOnce = False
runOnce1 = False




def log_fire_event(file_name, location, description):
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook.active
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    next_id = sheet.max_row + 1
    sheet.append([next_id, timestamp, location, description])
    workbook.save(file_name)

def initialize_excel(file_name):
    try:
        workbook = openpyxl.load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Fire Events'
        sheet.append(['ID', 'Timestamp', 'Location', 'Description'])
        workbook.save(file_name)

def play_alarm_sound_function():
    global runOnce
    pygame.mixer.music.load('fire_alarm.mp3')
    pygame.mixer.music.play()
    while pygame.mixer.music.get_busy():
        continue
    print("Fire alarm end")
    runOnce = False

def send_mail_function():
    global runOnce1
    if runOnce1:
        return

    EMAIL_ADDRESS = 'arsh.jain2004@gmail.com'
    EMAIL_PASSWORD = 'rfeelsfmzwsaddqb'
    email_list = ['arsh.jain2003@gmail.com', 'sakshamsabharwal2@gmail.com']  

    body = """
    <html>
    <head></head>
    <body>
        <h1>Fire has initiated in the building.</h1>
        <p>Please remember the following key guidelines:</p>
        <ul>
            <li>Stay Low: If encountering smoke, stay low to the ground to avoid inhaling harmful fumes.</li>
            <li>Use Exit Routes: Utilize the nearest exit or fire escape; avoid using elevators during a fire.</li>
            <li>Check Doors: Before opening any doors, check if they're hot. If hot, do not open—fire may be on the other side.</li>
            <li>Move to Safety: Once outside, move to a safe distance away from the building and call emergency services.</li>
            <li>Take Shelter if Necessary: If unable to exit safely, find a room with a window, seal cracks around doors to block smoke, and signal for help.</li>
        </ul>
        <p>It's vital for everyone's safety that these precautions are followed. Please review and familiarize yourself with the building's evacuation plan.</p>
        <p>Stay safe and vigilant,</p>
        <p><strong>Arsh Jain</strong></p>
        <p>Fire Safety Department</p>
    </body>
    </html>
    """

    context = ssl.create_default_context()

    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
            smtp.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            for email in email_list:
                msg = EmailMessage()
                msg['Subject'] = 'Warning - Fire Initiated'
                msg['From'] = EMAIL_ADDRESS
                msg['To'] = email
                msg.set_content('')
                msg.add_alternative(body, subtype="html")
                smtp.send_message(msg)
                print(f"Email sent to {email}")

        runOnce1 = True
    except Exception as e:
        print(f"Failed to send email: {e}")


pygame.mixer.init()
while True:
    ret, frame = vid.read()
    if not ret:
        break

    fire = fire_cascade.detectMultiScale(frame, 1.1, 5)
    initialize_excel('fire_detection.xlsx')

    fire_detected = False

    for (x, y, w, h) in fire:
        detected_region = frame[y:y+h, x:x+w]
        
        cv2.rectangle(frame, (x-20, y-20), (x+w+20, y+h+20), (255, 0, 0), 2)
        fire_detected = True

        if not runOnce:
            print("Fire alarm initiated")
            threading.Thread(target=play_alarm_sound_function).start()
            runOnce = True

        if not runOnce1:
            print("Mail sent initiated")
            threading.Thread(target=send_mail_function).start()
            threading.Thread(target=sendsms).start()
            threading.Thread(target=call).start()
            log_fire_event('fire_detection.xlsx', 'Warehouse 1', 'Reason by the dept.')
            runOnce1 = True

    cv2.imshow('frame', frame)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break


vid.release()
cv2.destroyAllWindows()




